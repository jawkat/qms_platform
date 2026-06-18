import os
from pathlib import Path
from flask import current_app


ALLOWED_EXTENSIONS = {'.zip', '.csv', '.xlsx', '.xls'}
MAX_FILE_SIZE_MB = 100


def validate_import_file(file_path):
    """
    Validate import file exists, has allowed extension, and size is reasonable.
    Returns (is_valid, error_messages).
    """
    errors = []

    if not file_path:
        errors.append('File path is required.')
        return False, errors

    if not isinstance(file_path, (str, Path)):
        errors.append(f'Invalid file path type: {type(file_path)}')
        return False, errors

    file_path = Path(file_path)

    if not file_path.exists():
        errors.append(f'File does not exist: {file_path}')
        return False, errors

    if file_path.suffix.lower() not in ALLOWED_EXTENSIONS:
        errors.append(f'File extension {file_path.suffix} not allowed. Allowed: {ALLOWED_EXTENSIONS}')
        return False, errors

    file_size_mb = file_path.stat().st_size / (1024 * 1024)
    if file_size_mb > MAX_FILE_SIZE_MB:
        errors.append(f'File size {file_size_mb:.1f}MB exceeds max {MAX_FILE_SIZE_MB}MB')
        return False, errors

    return True, errors


def run_mass_import(file_path, options=None):
    """
    Orchestrate mass import: validate file, then delegate to handler based on type.
    Returns stats dict with imported count, validated count, errors list.
    """
    options = options or {}
    stats = {
        'file_path': str(file_path),
        'imported': 0,
        'validated': 0,
        'errors': [],
        'validation_failed': 0,
    }

    is_valid, validation_errors = validate_import_file(file_path)
    if not is_valid:
        stats['validation_failed'] = len(validation_errors)
        stats['errors'].extend(validation_errors)
        return stats

    file_path = Path(file_path)
    ext = file_path.suffix.lower()

    try:
        if ext == '.zip':
            result = _import_from_zip(file_path, options)
        elif ext == '.csv':
            result = _import_from_csv(file_path, options)
        elif ext in {'.xlsx', '.xls'}:
            result = _import_from_excel(file_path, options)
        else:
            stats['errors'].append(f'No handler for extension {ext}')
            return stats

        stats.update(result)
    except Exception as e:
        stats['errors'].append(f'Import failed: {str(e)}')
        current_app.logger.exception(f'Import error for {file_path}')

    return stats


def _import_from_zip(file_path, options):
    """Import from ZIP archive (e.g. directory of files or structured dataset)."""
    import zipfile

    stats = {'imported': 0, 'validated': 0, 'errors': []}

    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            files_in_zip = zf.namelist()
            # Placeholder: validate structure, extract, and import each file
            stats['validated'] = len(files_in_zip)
            # In real implementation, extract and process each CSV/Excel sheet
            stats['imported'] = len(files_in_zip)
    except Exception as e:
        stats['errors'].append(f'ZIP extraction error: {e}')

    return stats


def _import_from_csv(file_path, options):
    """Import from CSV file."""
    import csv

    stats = {'imported': 0, 'validated': 0, 'errors': []}

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):  # start at 2 (after header)
                if not row or not any(row.values()):
                    continue
                # Validate row structure and persist
                # Placeholder validation: check required fields
                stats['validated'] += 1
                stats['imported'] += 1
    except Exception as e:
        stats['errors'].append(f'CSV parsing error: {e}')

    return stats


def _import_from_excel(file_path, options):
    """Import from Excel file."""
    try:
        import openpyxl
    except ImportError:
        return {'imported': 0, 'validated': 0, 'errors': ['openpyxl not installed']}

    stats = {'imported': 0, 'validated': 0, 'errors': []}

    try:
        wb = openpyxl.load_workbook(file_path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                # Validate row and persist
                stats['validated'] += 1
                stats['imported'] += 1
    except Exception as e:
        stats['errors'].append(f'Excel parsing error: {e}')

    return stats
