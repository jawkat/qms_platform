import os
from pathlib import Path
from flask import current_app


def validate_query_params(query_params):
    """
    Validate export query parameters (filters, columns, format).
    Returns (is_valid, error_messages).
    """
    errors = []

    if not isinstance(query_params, dict):
        errors.append(f'query_params must be dict, got {type(query_params)}')
        return False, errors

    # Check required keys
    if 'entity' not in query_params:
        errors.append('entity (e.g. "actions", "preuves") is required')
        return False, errors

    allowed_entities = {'actions', 'preuves', 'entreprises', 'utilisateurs', 'textes'}
    if query_params['entity'] not in allowed_entities:
        errors.append(f'entity must be one of {allowed_entities}')
        return False, errors

    # Validate filters if present
    filters = query_params.get('filters', {})
    if not isinstance(filters, dict):
        errors.append('filters must be a dict')
        return False, errors

    # Validate columns if present
    columns = query_params.get('columns', [])
    if not isinstance(columns, list):
        errors.append('columns must be a list')
        return False, errors

    return True, errors


def run_excel_export(query_params, output_path=None):
    """
    Validate query_params and execute Excel export.
    Returns stats dict with row_count, file_path, errors.
    """
    stats = {
        'entity': query_params.get('entity', 'unknown'),
        'row_count': 0,
        'file_path': None,
        'errors': [],
        'validation_failed': 0,
    }

    is_valid, validation_errors = validate_query_params(query_params)
    if not is_valid:
        stats['validation_failed'] = len(validation_errors)
        stats['errors'].extend(validation_errors)
        return stats

    try:
        # Default output path if not provided
        if output_path is None:
            output_dir = Path(current_app.config.get('UPLOAD_FOLDER', '/tmp'))
            output_dir.mkdir(parents=True, exist_ok=True)
            entity = query_params.get('entity', 'export')
            output_path = output_dir / f'{entity}_export.xlsx'

        output_path = Path(output_path)

        # Execute export logic
        result = _export_to_excel(query_params, output_path)
        stats.update(result)

    except Exception as e:
        stats['errors'].append(f'Export failed: {str(e)}')
        current_app.logger.exception(f'Export error for {query_params}')

    return stats


def _export_to_excel(query_params, output_path):
    """
    Execute export: fetch data, validate it, and write to Excel.
    Returns dict with row_count, file_path, errors.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return {
            'row_count': 0,
            'file_path': None,
            'errors': ['openpyxl not installed. Install with: pip install openpyxl']
        }

    stats = {
        'row_count': 0,
        'file_path': str(output_path),
        'errors': [],
    }

    try:
        entity = query_params.get('entity')
        filters = query_params.get('filters', {})
        columns = query_params.get('columns', [])

        # Placeholder: fetch data from DB based on entity and filters
        # In real implementation, query DB using filters
        data_rows = _fetch_export_data(entity, filters)

        if not data_rows:
            stats['errors'].append(f'No data found for entity={entity} with filters={filters}')
            return stats

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = entity

        # Write headers (use provided columns or default to data keys)
        if columns:
            headers = columns
        else:
            headers = list(data_rows[0].keys()) if data_rows else []

        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font

        # Write data rows
        for row_idx, row_data in enumerate(data_rows, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, '')

        stats['row_count'] = len(data_rows)

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

    except Exception as e:
        stats['errors'].append(f'Excel writing error: {e}')

    return stats


def _fetch_export_data(entity, filters):
    """
    Placeholder: fetch data from database based on entity and filters.
    In real implementation, build SQLAlchemy query and execute.
    """
    # Example: would return list of dicts
    return [
        {'id': 1, 'name': 'Action 1', 'status': 'En cours'},
        {'id': 2, 'name': 'Action 2', 'status': 'Complétée'},
    ]


def run_evaluation_export(texte_id, entreprise_id, output_path):
    """Generate Excel export for evaluations of a texte. Returns stats dict."""
    import io
    import re
    import html as html_lib
    from pathlib import Path
    from sqlalchemy import and_
    from sqlalchemy.orm import selectinload
    from app import db
    from app.models import EvaluationArticle, EntrepriseTexte, Article, ActionCorrective

    stats = {
        'texte_id': texte_id,
        'row_count': 0,
        'action_count': 0,
        'file_path': str(output_path),
        'errors': [],
    }

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        stats['errors'].append('openpyxl is required for Excel export')
        return stats

    try:
        output_path = Path(output_path)

        entreprise_texte = db.session.query(EntrepriseTexte).filter_by(
            id=texte_id, entreprise_id=entreprise_id
        ).first()
        if not entreprise_texte:
            stats['errors'].append('EntrepriseTexte not found')
            return stats

        def _strip_html(text):
            if not text:
                return ''
            cleaned = html_lib.unescape(str(text))
            cleaned = re.sub(r'<[^>]+>', '', cleaned)
            cleaned = re.sub(r'\s+', ' ', cleaned).strip()
            return cleaned

        articles_with_evaluations = (
            db.session.query(Article, EvaluationArticle)
            .outerjoin(
                EvaluationArticle,
                and_(
                    EvaluationArticle.article_id == Article.id,
                    EvaluationArticle.entreprise_texte_id == texte_id,
                ),
            )
            .filter(Article.texte_version_id == entreprise_texte.texte_version_id)
            .options(
                selectinload(Article.evaluations),
                selectinload(EvaluationArticle.justificatifs),
                selectinload(EvaluationArticle.actions),
            )
            .order_by(Article.numero_article)
            .all()
        )

        workbook = Workbook()
        worksheet = workbook.active
        texte_title = (entreprise_texte.version.texte.titre or 'texte').strip()
        clean_title = re.sub(r'[?/\\*\[\]:]', '', texte_title)
        worksheet.title = (clean_title[:31] or 'evaluations')

        worksheet.append([f"Texte: {texte_title}"])
        worksheet.append(['Contenu article', 'Applicabilite', 'Conformite', 'Preuve', 'Action', 'Commentaire'])

        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E5C6E', end_color='2E5C6E', fill_type='solid')
        border_side = Side(style='thin', color='D0D0D0')
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        worksheet.merge_cells('A1:F1')
        title_cell = worksheet['A1']
        title_cell.font = title_font
        title_cell.alignment = Alignment(vertical='center')

        for col in range(1, 7):
            cell = worksheet.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        for article, evaluation in articles_with_evaluations:
            article_label = _strip_html(article.contenu if article else '')
            applicable_label = evaluation.applicable.value if evaluation and evaluation.applicable else ''
            conforme_label = evaluation.conforme.value if evaluation and evaluation.conforme else ''
            preuves = ''
            actions = ''
            commentaire = ''
            if evaluation:
                preuves = ', '.join(
                    sorted({doc.nom_fichier for doc in evaluation.justificatifs if doc.nom_fichier})
                )
                actions = ', '.join(
                    [action.description for action in evaluation.actions if action.description]
                )
                commentaire = _strip_html(evaluation.observation or '')
            worksheet.append([article_label, applicable_label, conforme_label, preuves, actions, commentaire])

        worksheet.freeze_panes = 'A3'
        worksheet.column_dimensions['A'].width = 80
        worksheet.column_dimensions['B'].width = 16
        worksheet.column_dimensions['C'].width = 16
        worksheet.column_dimensions['D'].width = 30
        worksheet.column_dimensions['E'].width = 40
        worksheet.column_dimensions['F'].width = 40

        for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = cell_border
                if cell.column in (1, 4, 5, 6):
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='top')

        actions_sheet = workbook.create_sheet('Actions')
        actions_sheet.append([f"Texte: {texte_title}"])
        actions_sheet.append([
            'Article', 'Action', 'Cause identifiee', 'Responsable',
            'Date echeance', 'Date realisation', 'Statut', 'Efficacite', 'Preuve',
        ])

        actions_sheet.merge_cells('A1:I1')
        actions_title_cell = actions_sheet['A1']
        actions_title_cell.font = title_font
        actions_title_cell.alignment = Alignment(vertical='center')

        for col in range(1, 10):
            cell = actions_sheet.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = cell_border

        action_rows = (
            ActionCorrective.query
            .join(EvaluationArticle, ActionCorrective.evaluation_article_id == EvaluationArticle.id)
            .join(Article, EvaluationArticle.article_id == Article.id)
            .filter(EvaluationArticle.entreprise_texte_id == texte_id)
            .filter(Article.texte_version_id == entreprise_texte.texte_version_id)
            .options(selectinload(ActionCorrective.responsable))
            .order_by(Article.numero_article, ActionCorrective.id)
            .all()
        )

        for action in action_rows:
            article_label = ''
            if action.evaluation and action.evaluation.article:
                article_label = _strip_html(action.evaluation.article.contenu or '')
            responsable_label = ''
            if action.responsable:
                responsable_label = f"{action.responsable.prenom} {action.responsable.nom}".strip()
            actions_sheet.append([
                article_label,
                _strip_html(action.description or ''),
                _strip_html(action.cause_identifiee or ''),
                responsable_label,
                action.date_echeance.isoformat() if action.date_echeance else '',
                action.date_realisation.isoformat() if action.date_realisation else '',
                action.statut or '',
                action.efficacite if action.efficacite is not None else '',
                action.preuve or '',
            ])

        actions_sheet.freeze_panes = 'A3'
        actions_sheet.column_dimensions['A'].width = 60
        actions_sheet.column_dimensions['B'].width = 40
        actions_sheet.column_dimensions['C'].width = 40
        actions_sheet.column_dimensions['D'].width = 25
        actions_sheet.column_dimensions['E'].width = 14
        actions_sheet.column_dimensions['F'].width = 16
        actions_sheet.column_dimensions['G'].width = 16
        actions_sheet.column_dimensions['H'].width = 12
        actions_sheet.column_dimensions['I'].width = 20

        for row in actions_sheet.iter_rows(min_row=3, max_row=actions_sheet.max_row, min_col=1, max_col=9):
            for cell in row:
                cell.border = cell_border
                if cell.column in (1, 2, 3, 4, 9):
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='top')

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_path))
        stats['row_count'] = len(articles_with_evaluations)
        stats['action_count'] = len(action_rows)

    except Exception as e:
        stats['errors'].append(str(e))

    return stats
