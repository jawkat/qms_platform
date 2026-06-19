import io
from datetime import date, datetime
from flask import Response, current_app
from app import db
from app.models.actions import ActionCorrective
from app.models.audit import Audit
from app.models.nonconformite import NonConformite
from app.models.qualite import Risque


def export_actions_excel(entreprise_id, domaine=None):
    query = ActionCorrective.query.filter_by(entreprise_id=entreprise_id)
    if domaine:
        query = query.filter_by(domaine=domaine)
    items = query.order_by(ActionCorrective.date_creation.desc()).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Actions'

    headers = ['ID', 'Description', 'Domaine', 'Type', 'Priorité', 'Statut',
               'Responsable', 'Échéance', 'Date réalisation', 'Cause', 'RCA 1', 'RCA 2']
    header_fill = PatternFill(start_color='2E5C6E', end_color='2E5C6E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_idx, a in enumerate(items, 2):
        resp = f'{a.responsable.prenom} {a.responsable.nom}' if a.responsable else ''
        ws.cell(row=row_idx, column=1, value=a.id)
        ws.cell(row=row_idx, column=2, value=a.description or '')
        ws.cell(row=row_idx, column=3, value=a.domaine or '')
        ws.cell(row=row_idx, column=4, value=a.type_action or '')
        ws.cell(row=row_idx, column=5, value=a.priorite or '')
        ws.cell(row=row_idx, column=6, value=a.statut or '')
        ws.cell(row=row_idx, column=7, value=resp)
        ws.cell(row=row_idx, column=8, value=a.date_echeance.isoformat() if a.date_echeance else '')
        ws.cell(row=row_idx, column=9, value=a.date_realisation.isoformat() if a.date_realisation else '')
        ws.cell(row=row_idx, column=10, value=a.cause_identifiee or '')
        ws.cell(row=row_idx, column=11, value=a.rca_pourquoi1 or '')
        ws.cell(row=row_idx, column=12, value=a.rca_pourquoi2 or '')

    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['G'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=actions_{date.today().isoformat()}.xlsx'}
    )


def export_audits_excel(entreprise_id, domaine=None):
    query = Audit.query.filter_by(entreprise_id=entreprise_id)
    if domaine:
        query = query.filter_by(domaine=domaine)
    items = query.order_by(Audit.date_creation.desc()).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Audits'

    headers = ['ID', 'Type', 'Domaine', 'Date', 'Auditeur', 'Résultat', 'Commentaire']
    header_fill = PatternFill(start_color='2E5C6E', end_color='2E5C6E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, a in enumerate(items, 2):
        auditeur = f'{a.auditeur.prenom} {a.auditeur.nom}' if a.auditeur else ''
        ws.cell(row=row_idx, column=1, value=a.id)
        ws.cell(row=row_idx, column=2, value=a.type or '')
        ws.cell(row=row_idx, column=3, value=a.domaine or '')
        ws.cell(row=row_idx, column=4, value=a.date_audit.isoformat() if a.date_audit else '')
        ws.cell(row=row_idx, column=5, value=auditeur)
        ws.cell(row=row_idx, column=6, value=a.resultat_global or '')
        ws.cell(row=row_idx, column=7, value=a.commentaire or '')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=audits_{date.today().isoformat()}.xlsx'}
    )


def export_nc_excel(entreprise_id, domaine=None):
    query = NonConformite.query.filter_by(entreprise_id=entreprise_id)
    if domaine:
        query = query.filter_by(domaine=domaine)
    items = query.order_by(NonConformite.date_creation.desc()).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Non-conformités'

    headers = ['ID', 'Réf', 'Domaine', 'Date', 'Description', 'Statut', 'Gravité', 'Responsable', 'Cause', 'Clôturée le']
    header_fill = PatternFill(start_color='2E5C6E', end_color='2E5C6E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, nc in enumerate(items, 2):
        resp = f'{nc.responsable.prenom} {nc.responsable.nom}' if nc.responsable else ''
        ws.cell(row=row_idx, column=1, value=nc.id)
        ws.cell(row=row_idx, column=2, value=nc.reference or '')
        ws.cell(row=row_idx, column=3, value=nc.domaine or '')
        ws.cell(row=row_idx, column=4, value=nc.date_nc.isoformat() if nc.date_nc else '')
        ws.cell(row=row_idx, column=5, value=nc.description or '')
        ws.cell(row=row_idx, column=6, value=nc.statut or '')
        ws.cell(row=row_idx, column=7, value=nc.gravite or '')
        ws.cell(row=row_idx, column=8, value=resp)
        ws.cell(row=row_idx, column=9, value=nc.cause or '')
        ws.cell(row=row_idx, column=10, value=nc.cloture_le.isoformat() if nc.cloture_le else '')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=nc_{date.today().isoformat()}.xlsx'}
    )


def export_risques_excel(entreprise_id):
    items = Risque.query.filter_by(entreprise_id=entreprise_id)\
        .order_by(Risque.date_creation.desc()).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Risques'

    headers = ['ID', 'Désignation', 'Cause', 'Effet', 'Gravité', 'Probabilité', 'Détection', 'IPR', 'Niveau', 'Statut', 'Processus', 'Plan traitement']
    header_fill = PatternFill(start_color='2E5C6E', end_color='2E5C6E', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, r in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=r.id)
        ws.cell(row=row_idx, column=2, value=r.designation or '')
        ws.cell(row=row_idx, column=3, value=r.cause or '')
        ws.cell(row=row_idx, column=4, value=r.effet or '')
        ws.cell(row=row_idx, column=5, value=r.gravite)
        ws.cell(row=row_idx, column=6, value=r.probabilite)
        ws.cell(row=row_idx, column=7, value=r.detection)
        ws.cell(row=row_idx, column=8, value=r.ipr)
        ws.cell(row=row_idx, column=9, value=r.niveau_risque)
        ws.cell(row=row_idx, column=10, value=r.statut or '')
        ws.cell(row=row_idx, column=11, value=r.processus_concerne or '')
        ws.cell(row=row_idx, column=12, value=r.plan_traitement or '')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=risques_{date.today().isoformat()}.xlsx'}
    )
